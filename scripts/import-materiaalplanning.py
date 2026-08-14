"""
Eenmalige import van de Excel-materiaalplanning naar het platform.

    python3 scripts/import-materiaalplanning.py            # dry run
    python3 scripts/import-materiaalplanning.py --apply    # schrijft weg

Leest .env.local voor de Supabase-sleutels. Per maandblad staan de materialen
in de kolommen en de datums in de rijen; opeenvolgende dagen met dezelfde
persoon worden één reservatie (met dezelfde ophaal- en terugbrengtijd), maar
net als de app wordt er per bezette dag een rij weggeschreven.
"""
import sys, os, json, re, datetime, unicodedata, urllib.request, urllib.parse
from collections import defaultdict, Counter

import openpyxl

XLSX_PATH  = '/Users/alexandervandenbranden/Downloads/Materiaalplanning - SHG.xlsx'
FROM_DATE  = '2026-06-01'          # laatste twee maanden + alles daarna
APPLY      = '--apply' in sys.argv
BACKUP_DIR = '/private/tmp/claude-501/-Users-alexandervandenbranden/5b1b69b1-b664-41d8-b5bb-3fcdf585a3ae/scratchpad'

# Niet meer in dienst — hun reservaties slaan we over.
DEPARTED = {'jens', 'stijn', 'arnor', 'kobe', 'steven', 'ruben', 'charlotte', 'sander'}

# Labels die we in hun geheel overslaan. "Arne - EXTERN" is externe verhuur en
# hoort niet op naam van Arne te komen.
SKIP_LABELS = {'arne - extern'}

# Voornamen die niet (of verkeerd) matchen op de teamlijst. Jelle staat er als
# bijnaam bij omdat de lijst ook een andere Jelle bevat.
ALIASES  = {
    'kenny': 'Kenneth Staelens',
    'jarne': 'Jarne Wouters',
    'jelle': 'Jelle Desterbecq',
}


def norm(s):
    s = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()


# ── Supabase REST ────────────────────────────────────────────────────────────
def load_env():
    env = {}
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for line in open(os.path.join(root, '.env.local')):
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


ENV = load_env()
URL = ENV['NEXT_PUBLIC_SUPABASE_URL']
KEY = ENV['SUPABASE_SERVICE_ROLE_KEY']


def api(path, method='GET', body=None, extra_headers=None):
    req = urllib.request.Request(f'{URL}{path}', method=method)
    req.add_header('apikey', KEY)
    req.add_header('Authorization', f'Bearer {KEY}')
    req.add_header('Content-Type', 'application/json')
    for k, v in (extra_headers or {}).items():
        req.add_header(k, v)
    data = json.dumps(body).encode() if body is not None else None
    with urllib.request.urlopen(req, data) as r:
        raw = r.read()
        return json.loads(raw) if raw else None


# ── Roster ───────────────────────────────────────────────────────────────────
def build_roster():
    first = {}
    clients = api('/rest/v1/clients?select=id&name=eq.Sporthouse')
    contacts = api(f'/rest/v1/contacts?select=name&client_id=eq.{clients[0]["id"]}')
    users = api('/auth/v1/admin/users?per_page=200')

    def add(full):
        if not full:
            return
        key = norm(full.split(' ')[0])
        first.setdefault(key, full)

    for c in contacts:
        add(c['name'])
    for u in users.get('users', []):
        if u.get('app_metadata', {}).get('freelancer'):
            continue
        m = u.get('user_metadata') or {}
        add(m.get('full_name') or m.get('name'))

    first.update(ALIASES)
    return first


def resolve(raw, roster):
    """None = vrij, ('skip', label) = vertrokken, (naam, matched, label) anders."""
    label = ' '.join(str(raw).split())
    if not label or norm(label) == 'vrij':
        return None
    if norm(label) in SKIP_LABELS:
        return ('skip', label)
    head = norm(re.split(r'[\s\-–—:,&]', label)[0])
    if head in DEPARTED:
        return ('skip', label)
    matched = roster.get(head)
    # Geen match is meestal geen persoon maar een partij (FoS, De Spor, KAPOT…).
    # Die nemen we letterlijk over: het materiaal is dan wel degelijk bezet.
    return (matched or label, matched is not None, label)


# ── Excel inlezen ────────────────────────────────────────────────────────────
roster = build_roster()
equipment = api('/rest/v1/equipment?select=id,name,category')
by_name = defaultdict(list)
for e in equipment:
    by_name[norm(e['name'])].append(e)


def lookup_equipment(header, occurrence):
    k = norm(header)
    if k in by_name:
        return by_name[k][0]
    if f'{k} {occurrence}' in by_name:            # "GoPro Harnas" -> "... 1"/"2"
        return by_name[f'{k} {occurrence}'][0]
    for key, lst in by_name.items():
        if key.startswith(k) or k.startswith(key):
            return lst[0]
    return None


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
occupancy   = defaultdict(set)     # (equipment_id, naam) -> {datums}
day_projects = {}
unknown_cols = Counter()
skipped      = Counter()
kept         = Counter()
cell_count   = 0

for sheet in wb.sheetnames:
    if sheet.startswith('FOS -'):
        continue
    ws = wb[sheet]

    seen, columns = Counter(), []
    for c in range(3, ws.max_column + 1):
        raw = ws.cell(2, c).value
        if not raw or not str(raw).strip():
            continue
        header = ' '.join(str(raw).split())
        seen[norm(header)] += 1
        item = lookup_equipment(header, seen[norm(header)])
        if item is None:
            unknown_cols[header] += 1
            continue
        columns.append((c, item))

    for r in range(3, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime):
            continue
        iso = d.date().isoformat()
        if iso < FROM_DATE:
            continue

        proj = ws.cell(r, 2).value
        if proj and str(proj).strip():
            day_projects[iso] = ' '.join(str(proj).split())

        for c, item in columns:
            res = resolve(ws.cell(r, c).value or '', roster)
            if res is None:
                continue
            if res[0] == 'skip':
                skipped[res[1]] += 1
                continue
            name, _matched, label = res
            cell_count += 1
            kept[label] += 1
            occupancy[(item['id'], name)].add(iso)

# ── Opeenvolgende dagen -> runs (één ophaal/terugbreng-paar per run) ─────────
def add_days(iso, n):
    return (datetime.date.fromisoformat(iso) + datetime.timedelta(days=n)).isoformat()


rows, runs = [], 0
for (equipment_id, name), dates in occupancy.items():
    dates = sorted(dates)
    start = prev = dates[0]

    def flush(start, end):
        global runs
        runs += 1
        # 09:00 ophalen, 17:00 terugbrengen: na 10:30, dus de laatste dag telt
        # als volledig bezet — precies wat in Excel staat.
        pickup = f'{start}T09:00:00+02:00'
        ret    = f'{end}T17:00:00+02:00'
        d = start
        while d <= end:
            rows.append({
                'equipment_id': equipment_id,
                'reserved_by': name,
                'date': d,
                'pickup_datetime': pickup,
                'return_datetime': ret,
            })
            d = add_days(d, 1)

    for d in dates[1:]:
        if d == add_days(prev, 1):
            prev = d
            continue
        flush(start, prev)
        start = prev = d
    flush(start, prev)

# ── Rapport ──────────────────────────────────────────────────────────────────
all_dates = sorted({r['date'] for r in rows})
print(f"\n{'IMPORT' if APPLY else 'DRY RUN'} — vanaf {FROM_DATE}\n")
print(f'  bezette cellen      {cell_count}')
print(f'  reservaties (runs)  {runs}')
print(f'  rijen (1 per dag)   {len(rows)}')
print(f'  dagprojecten        {len(day_projects)}')
if all_dates:
    print(f'  periode             {all_dates[0]} .. {all_dates[-1]}')

print('\n  toegewezen aan:')
for label, n in kept.most_common():
    res = resolve(label, roster)
    tag = '' if res[1] else '   (letterlijk overgenomen)'
    print(f'    {n:4}  {label:24} -> {res[0]}{tag}')

if skipped:
    print('\n  overgeslagen:')
    for label, n in skipped.most_common():
        print(f'    {n:4}  {label}')

if unknown_cols:
    print('\n  kolommen zonder materiaal in het platform:')
    for h, n in unknown_cols.most_common():
        print(f'    {n:3}x  {h}')

if not APPLY:
    print('\nNiets geschreven. Voeg --apply toe om door te voeren.\n')
    sys.exit(0)

# ── Wegschrijven ─────────────────────────────────────────────────────────────
existing = api('/rest/v1/equipment_reservations?select=*')
existing_projects = api('/rest/v1/day_projects?select=*')
stamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
backup = os.path.join(BACKUP_DIR, f'materiaalplanning-backup-{stamp}.json')
with open(backup, 'w') as f:
    json.dump({'reservations': existing, 'day_projects': existing_projects}, f, indent=2)
print(f'\nback-up van {len(existing)} reservaties en {len(existing_projects)} dagprojecten -> {backup}')

api('/rest/v1/equipment_reservations?id=neq.00000000-0000-0000-0000-000000000000',
    method='DELETE', extra_headers={'Prefer': 'return=minimal'})
print('bestaande reservaties verwijderd')

created = 0
for i in range(0, len(rows), 500):
    chunk = rows[i:i + 500]
    api('/rest/v1/equipment_reservations', method='POST', body=chunk,
        extra_headers={'Prefer': 'return=minimal'})
    created += len(chunk)
    print(f'  {created}/{len(rows)} rijen')

if day_projects:
    api('/rest/v1/day_projects?on_conflict=date', method='POST',
        body=[{'date': d, 'project_name': p} for d, p in day_projects.items()],
        extra_headers={'Prefer': 'return=minimal,resolution=merge-duplicates'})
    print(f'dagprojecten: {len(day_projects)}')

print('\nKlaar.\n')
